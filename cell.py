import sys
import os
from openpyxl import load_workbook
from openpyxl import Workbook

rootdir = sys.path[0]

wb=Workbook()
sheet = wb.active

for parent,dirnames,filenames in os.walk(rootdir):
     for filename in filenames:                        
          if filename.find(".xlsx") >= 0:
               xfile = parent+"/"+filename
               wb2=load_workbook(xfile)
               ws2=wb2.active
               for row in ws2.rows:
                    ar = []
                    for cell in row:
                         ar.append(cell.value)
                    sheet.append(ar)

wb.save('all.xlsx')
